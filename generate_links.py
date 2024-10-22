# 按非路径生成关系
import pandas as pd
from openpyxl import load_workbook
from py2neo import Node, Graph, NodeMatcher, RelationshipMatcher, Relationship
from function_kc4 import eliminate_KC_ID
from function_adjacent2 import find_adjacent_duplicates
from collections import Counter

FILE = 'HRB_151318-7d'
SHEET = 0
FREQ = 400

# link and clean the old node
graph = Graph('bolt://localhost:7687', auth=('neo4j', 'duan9595'))
graph.delete_all()

# load Excel and read sheet
FileName = 'xls/%s.xlsx' % FILE
workbook = load_workbook(FileName)

# Loading the datasets
CheckInData = pd.read_excel(FileName,
                            sheet_name=workbook.sheetnames[SHEET],
                            dtype={'business_id': str})


business_ids = CheckInData['business_id'].to_list()
unique_business_ids = find_adjacent_duplicates(business_ids)

relate_tuple_list = [(unique_business_ids[index], unique_business_ids[index + 1])
                     for index in range(len(unique_business_ids) - 1)]

counter = dict(Counter(relate_tuple_list))
