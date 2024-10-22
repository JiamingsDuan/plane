"""划分基础路径"""
import pandas as pd
from openpyxl import load_workbook

# load xlsx and read sheet
FileName = 'xls/HRB_151318.xlsx'
workbook = load_workbook(FileName)
sheets = workbook.sheetnames
print(sheets)

# Loading the datasets
SheetName = sheets[0]
CheckInData = pd.read_excel(FileName, sheet_name=SheetName)
print(CheckInData.shape[0], CheckInData.shape[1])

# read sheet
RecordFrame = pd.DataFrame(columns=['business_id', 'business_name', 'business_time', 'business_label'])
print(type(RecordFrame))

# initialize the starting node and ending node
start_list = ['042E0102', '042E0105', '042E0115',
              '042E0200', '042E0201', '042E0211',
              '042E0212', '042E0219', '162R0300',
              '16200100', '16200101', '16240100', '162R0101']
end_list = ['042E0117', '042E0223', '01010915',
            '01010916', '01010917', '01010903',
            '042E0224', '01010600', 'nav']

print(len(start_list), len(end_list))


# mark "1","0","-1"
for index in range(CheckInData.shape[0]):

    Record = CheckInData.iloc[index, :]
    Record_Last = CheckInData.iloc[index - 1, :]

    if Record['business_id'] in start_list:
        Record['business_label'] = 1

    elif Record['business_id'] in end_list:
        Record['business_label'] = -1

    elif Record['business_id'] == '042E0200' and CheckInData.iloc[index + 1, :]['business_id'] == '042E0112':
        Record['business_label'] = 1

    elif Record['business_id'] == '042E0224' and Record_Last['business_id'] == '042E0200':
        Record['business_label'] = -1

    elif Record['business_id'] == '042E0200' and CheckInData.iloc[index + 1, :]['business_id'] != '042E0112' \
            and CheckInData.iloc[index + 1, :]['business_id'] != '042E0224':
        Record['business_label'] = 1
    else:
        Record['business_label'] = 0

    # print(Record.to_dict())
    RecordFrame.loc[len(RecordFrame)] = Record.tolist()

# print(RecordFrame)

# save to local
RecordFrame.to_csv('./csv/' + SheetName + '_new.csv', index=False, encoding='utf-8')
