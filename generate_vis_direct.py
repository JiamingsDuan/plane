import pandas as pd
from openpyxl import load_workbook
# from py2neo import Node
# from py2neo import Graph
# from py2neo import NodeMatcher
# from py2neo import RelationshipMatcher
# from py2neo import Relationship
from pyvis.network import Network
from collections import Counter
from function_kc1 import generate_hex_colors
from function_adjacent2 import find_adjacent_duplicates

# parameter
FILE = '0726'
FREQ = 1
SHEET = 0

# load Excel and read sheet
FileName = 'xls/%s.xlsx' % FILE
workbook = load_workbook(FileName)

# Loading the datasets
CheckInData = pd.read_excel(FileName,
                            sheet_name=workbook.sheetnames[SHEET],
                            dtype={'business_id': str})

# 统计business_id种类数量
counter = dict(Counter(CheckInData.loc[:, 'business_id'].to_list()))

# 创建一个Network对象
net = Network(
    notebook=False,
    directed=True,
    # select_menu=True
)
# net.show_buttons(filter_=['physics'])

# 添加节点
for uid, count in counter.items():
    net.add_node(uid, label=uid, color='grey', size=pow(count, 0.5))
    # typing_index = CheckInData[CheckInData['business_id'] == uid].index.to_list()[0]
    # typing_label = CheckInData.loc[typing_index, 'type_id']
    # if typing_label == 'W':
    #     net.add_node(uid, label=uid, color='#333399', size=pow(counter[uid], 0.5))
    # elif typing_label == 'R':
    #     net.add_node(uid, label=uid, color='#00CED1', size=pow(counter[uid], 0.5))
    # else:
    #     net.add_node(uid, label=uid, color='black', size=pow(counter[uid], 0.5))

business_ids = CheckInData['business_id'].to_list()
unique_business_ids = find_adjacent_duplicates(business_ids)

relate_tuple_list = [(unique_business_ids[index], unique_business_ids[index + 1])
                     for index in range(len(unique_business_ids) - 1)]

counter_tuple = dict(Counter(relate_tuple_list))

# 生成颜色编码表
link_color = generate_hex_colors(len(counter_tuple))
color_index = 0

for link, size in counter_tuple.items():
    # 根据节点属性查找节点
    name1 = link[0]
    name2 = link[-1]
    if counter_tuple[(name1, name2)] < FREQ:
        pass
    else:
        net.add_edge(name1,
                     name2,
                     value=size,
                     color=link_color[color_index],
                     title='频次%s' % size,
                     )
    color_index = color_index + 1

net.width = 2560
net.height = 1440

# 设置边的宽度与权重成正比
for edge in net.edges:
    edge['width'] = edge['value'] * 0.5

net.repulsion(180)

# 显示图形
net.show('htm/graph_%s.html' % FILE)
