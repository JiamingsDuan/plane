import pandas as pd
from openpyxl import load_workbook
from collections import Counter
from function_brt import determine_label

# important parameter
SHEET = 0
TIME_STEP = 3

# load Excel and read sheet
FileName = 'xls/HRB_151318.xlsx'
# obtain the sheets
workbook = load_workbook(FileName)

# Loading the datasets
CheckInData = pd.read_csv('csv/20240331_class.csv',
                          dtype={'business_id': str})

CheckInData['time_interval'] = ''
CheckInData['time_splitting'] = ''

ten_second_route_sub = [0, ]
# login index time_splitting is b
CheckInData.loc[0, 'time_splitting'] = 'b'
# second index to last index
for index in range(CheckInData.shape[0]):
    if index < CheckInData.shape[0] - 1:
        # generate time interval
        subtraction = pd.to_datetime(CheckInData.loc[index + 1, 'business_time']) \
                      - pd.to_datetime(CheckInData.loc[index, 'business_time'])
        spl = round(subtraction.total_seconds() % 60, 3)
        # generate 10 second node
        if spl >= TIME_STEP:
            CheckInData.loc[index + 1, 'time_splitting'] = 's'
        else:
            CheckInData.loc[index + 1, 'time_splitting'] = 'b'
        ten_second_route_sub.append(spl)
        # print(round(subtraction.total_seconds() % 60, 3))
    else:
        pass

# add time sub
CheckInData['time_interval'] = ten_second_route_sub
# save to local
# column = [x for x in CheckInData.columns if x not in ['business_name']]
# CheckInData.to_csv('./csv/' + workbook.sheetnames[SHEET] + '_time_split_route.csv', index=False, encoding='utf-8')
# obtain node marking
splitting_index = CheckInData.loc[CheckInData['time_splitting'] == 's'].index.tolist()

# add route label(basic label)
ten_second_route_labels = []
ten_second_route_bools = []
for ind in range(len(splitting_index)):
    if ind < len(splitting_index) - 1:
        ten_second_route_node1 = CheckInData.loc[splitting_index[ind], 'business_id']
        ten_second_route_node2 = CheckInData.loc[splitting_index[ind + 1] - 1, 'business_id']
        ten_second_route_label = determine_label(starting=ten_second_route_node1, ending=ten_second_route_node2)
        if ten_second_route_label == '-':
            ten_second_route_bools.append('False')
        else:
            ten_second_route_bools.append('True')
        ten_second_route_labels.append(ten_second_route_label)
        # print(ten_second_route_node1, ten_second_route_node2, ten_second_route_label)
    else:
        pass

frame = pd.DataFrame(columns=['序号', '路径内业务ID', '距上一ID时间间隔', '符合', '路径标签（名称）'])

for ind in range(len(splitting_index)):

    if ind < len(splitting_index) - 1:
        BusinessID = CheckInData.loc[splitting_index[ind]:splitting_index[ind + 1] - 1, 'business_id'].to_list()
        # BusinessName = CheckInData.loc[splitting_index[ind]:splitting_index[ind + 1] - 1, 'business_name'].to_list()
        BusinessTimeSplit = \
            CheckInData.loc[splitting_index[ind]:splitting_index[ind + 1] - 1, 'time_interval'].to_list()

        IDFrame = '\n'.join(BusinessID)
        # NameFrame = '\n'.join(BusinessName)
        TimeSplit = '\n'.join([str(x) for x in BusinessTimeSplit])
        frame.loc[frame.shape[0], :] = [str(frame.shape[0] + 1), IDFrame, TimeSplit, '', '']

frame['符合'] = ten_second_route_bools
frame['路径标签（名称）'] = ten_second_route_labels
rate = int(Counter(ten_second_route_bools)['True']) / int(frame.shape[0])
print(rate)
frame.to_excel('./xls/' + workbook.sheetnames[SHEET] + '_' + str(TIME_STEP) + 'seconds.xlsx',
               index=False,
               encoding='utf-8')
