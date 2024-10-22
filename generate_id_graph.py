# 按非路径切分business_id作图
import pandas as pd
from openpyxl import load_workbook
from py2neo import Node, Graph, NodeMatcher, RelationshipMatcher, Relationship

from function_adjacent2 import find_adjacent_duplicates
from collections import Counter

FILE = 'HRB_151318'
SHEET = 0
FREQ = 80

# link and clean the old node
graph = Graph('bolt://localhost:7687', auth=('neo4j', 'duan9595'))
graph.delete_all()

# load Excel and read sheet
FileName = 'xls/%s.xlsx' % FILE
workbook = load_workbook(FileName)

# Loading the datasets
CheckInData = pd.read_excel(FileName,
                            sheet_name=workbook.sheetnames[SHEET],
                            dtype={'business_id': str})

# 统计business_id种类数量
counter = Counter(CheckInData.loc[:, 'business_id'].to_list())
print('amount of business_id', len(counter))
# 按点的数量初始化点
for i in range(len(counter)):
    node = Node(list(counter.keys())[i], name=list(counter.keys())[i])
    graph.create(node)


business_ids = CheckInData['business_id'].to_list()
unique_business_ids = find_adjacent_duplicates(business_ids)

relate_tuple_list = [(unique_business_ids[index], unique_business_ids[index + 1])
                     for index in range(len(unique_business_ids) - 1)]

node_matcher = NodeMatcher(graph)
relationship_matcher = RelationshipMatcher(graph)
counter_tuple = dict(Counter(relate_tuple_list))


for i in range(len(relate_tuple_list)):
    # 根据节点属性查找节点
    name1 = relate_tuple_list[i][0]
    name2 = relate_tuple_list[i][-1]
    # print(counter[(name1, name2)])
    node1 = node_matcher.match(name1).where(name=name1).first()
    node2 = node_matcher.match(name2).where(name=name2).first()
    if counter_tuple[(name1, name2)] < FREQ:
        pass
    else:
        if relationship_matcher.match((node1, node2), r_type=None).exists():
            pass
        else:
            relate = Relationship(node1, str(counter_tuple[(name1, name2)]), node2)
            graph.create(relate)
