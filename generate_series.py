import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from function_adjacent1 import find_adjacent_duplicates
from collections import Counter

# obtain the sheets

SHEET = 0
# load Excel and read sheet
FileName = '../xls/HRB_151318.xlsx'
workbook = load_workbook(FileName)

# Loading the datasets
CheckInData = pd.read_excel(FileName,
                            sheet_name=workbook.sheetnames[SHEET],
                            dtype={'business_id': str})


# Business_Time_Distance:mark "SUB"
subtractions = [0, ]
# second index to last index
for index in range(CheckInData.shape[0]):
    if index < CheckInData.shape[0] - 1:
        subtraction = pd.to_datetime(CheckInData.loc[index + 1, 'business_time']) \
                - pd.to_datetime(CheckInData.loc[index, 'business_time'])
        subtractions.append(round(subtraction.total_seconds() % 60, 3))
        # print(round(subtraction.total_seconds() % 60, 3))
    else:
        pass

picture = plt.figure(figsize=(16, 9))

ax1 = picture.add_subplot(4, 3, 1)
plt.title('first~600 BusinessTime--Period')
plt.xlim(0, len(subtractions[:601]) + 1)
plt.ylim(0, round(max(subtractions[:601]), 0) + 1)
plt.plot(subtractions,
         linestyle='-',
         linewidth=0.75,
         marker='*',
         markersize=0.6
         )

ax2 = picture.add_subplot(4, 3, 2)
plt.title('601~1200 BusinessTime--Period')
plt.xlim(601, 1201)
plt.ylim(0, round(max(subtractions[601:1201]), 0) + 1)
plt.plot(subtractions,
         linestyle='-',
         linewidth=0.75,
         marker='*',
         markersize=0.6
         )

ax3 = picture.add_subplot(4, 3, 3)
plt.title('1201~1800 BusinessTime--Period')
plt.xlim(1201, 1801)
plt.ylim(0, round(max(subtractions[1201:1801]), 0) + 1)
plt.plot(subtractions,
         linestyle='-',
         linewidth=0.75,
         marker='*',
         markersize=0.6
         )

ax4 = picture.add_subplot(4, 3, 4)
plt.title('1801~2400 BusinessTime--Period')
plt.xlim(1801, 2401)
plt.ylim(0, round(max(subtractions[1801:2401]), 0) + 1)
plt.plot(subtractions,
         linestyle='-',
         linewidth=0.75,
         marker='*',
         markersize=0.6
         )

ax5 = picture.add_subplot(4, 3, 5)
plt.title('2401~3000 BusinessTime--Period')
plt.xlim(2401, 3001)
plt.ylim(0, round(max(subtractions[2401:3001]), 0) + 1)
plt.plot(subtractions,
         linestyle='-',
         linewidth=0.75,
         marker='*',
         markersize=0.6
         )

ax6 = picture.add_subplot(4, 3, 6)
plt.title('3001~3600 BusinessTime--Period')
plt.xlim(3001, 3601)
plt.ylim(0, round(max(subtractions[3001:3601]), 0) + 1)
plt.plot(subtractions,
         linestyle='-',
         linewidth=0.75,
         marker='*',
         markersize=0.6,
         )

ax7 = picture.add_subplot(4, 3, 7)
plt.title('3601~4201 BusinessTime--Period')
plt.xlim(3601, 4200)
plt.ylim(0, round(max(subtractions[3601:4201]), 0) + 1)
plt.plot(subtractions,
         linestyle='-',
         linewidth=0.75,
         marker='*',
         markersize=0.6
         )

ax8 = picture.add_subplot(4, 3, 8)
plt.title('4201~4801 BusinessTime--Period')
plt.xlim(4201, 4800)
plt.ylim(0, round(max(subtractions[4201:4801]), 0) + 1)
plt.plot(subtractions,
         linestyle='-',
         linewidth=0.75,
         marker='*',
         markersize=0.6
         )

ax9 = picture.add_subplot(4, 3, 9)
plt.title('4801~5400 BusinessTime--Period')
plt.xlim(4801, 5400)
plt.ylim(0, round(max(subtractions[4801:5401]), 0) + 1)
plt.plot(subtractions,
         linestyle='-',
         linewidth=0.75,
         marker='*',
         markersize=0.6
         )

ax10 = picture.add_subplot(4, 3, 10)
plt.title('5401~6000 BusinessTime--Period')
plt.xlim(5401, 6000)
plt.ylim(0, round(max(subtractions[5401:6001]), 0) + 1)
plt.plot(subtractions,
         linestyle='-',
         linewidth=0.75,
         marker='*',
         markersize=0.6
         )

ax11 = picture.add_subplot(4, 3, 11)
plt.title('6001~last BusinessTime--Period')
plt.xlim(6001, len(subtractions) + 1)
plt.ylim(0, round(max(subtractions[6001:]), 0) + 1)
plt.plot(subtractions,
         linestyle='-',
         linewidth=0.75,
         marker='*',
         markersize=0.6
         )

# ax12 = picture.add_subplot(4, 3, 12)
# plt.title('6601~last BusinessTime--Period')
# plt.xlim(6601, len(subtractions) + 1)
# plt.ylim(0, round(max(subtractions[6601:]), 0) + 1)
# plt.plot(subtractions,
#          linestyle='-',
#          linewidth=0.75,
#          marker='*',
#          markersize=0.6
#          )

plt.savefig('fig/Period.pdf')
plt.show()

# # obtain repeated tuples
# repeat_index_tuple = find_adjacent_duplicates(list(CheckInData))
# print(len(repeat_index_tuple))
#
# # order the repeated index tuple
# dropped_index_list = []
# dropped_value_list = []
# for repeated in repeat_index_tuple:
#     dropped_index_list.append(repeated[-1])
#     dropped_value_list.append(list(CheckInData)[repeated[0]])
#
#
# print(len(dropped_index_list), len(dropped_value_list))
# CheckInData.drop(dropped_index_list, inplace=True)
#
#
# # Check the processed Series
# repeat_index_tuple_again = find_adjacent_duplicates(list(CheckInData))
# print(len(list(CheckInData)))

# RouteCounts = Counter(dropped_value_list).values()
# RouteLabels = Counter(dropped_value_list).keys()
