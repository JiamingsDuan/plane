import sqlite3
import pandas as pd
import os
from openpyxl import load_workbook

DIR = 'xls/'
# FILE = 'HRB_151318'


# load Excel and read sheet
xls_file_name_lst = os.listdir(DIR)
conn = sqlite3.connect('E:/BaiduNetdiskDownload/database/CheckIndata.db')
name = '0725(new)'
CheckInData_0 = pd.read_excel('xls/%s.xlsx' % name, dtype={'business_id': str})
print('saving file xls/%s.sql......' % name)
CheckInData_0.to_sql(name='travelsky' + name, con=conn, if_exists='replace', index=False)
# CheckInData_0.to_sql(name='travelsky', con=conn, if_exists='append', index=False)
conn.close()


# workbook = load_workbook('xls/%s.xlsx' % FILE).sheetnames

# for name in workbook:
#     CheckInData = pd.read_excel('xls/%s.xlsx' % FILE, sheet_name=name)
#     CheckInData.to_sql(name=name,
#                        con=conn,
#                        if_exists='replace',
#                        index=False)

# for filename in xls_file_name_lst:
#     name = filename.split('.')[0]
#     print('loading file xls/%s......' % filename)
#     CheckInData_0 = pd.read_excel('xls/%s.xlsx' % name, dtype={'business_id': str})
#     print('saving file xls/%s.sql......' % name)
#     CheckInData_0.to_sql(name=name,
#                          con=conn,
#                          if_exists='replace',
#                          index=False)
# CheckInData_0.to_sql(name='travelsky',
#                      con=conn,
#                      if_exists='append',
#                      index=False)
# conn.close()
