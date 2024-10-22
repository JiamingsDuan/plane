"""聚类"""
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from sklearn.cluster import KMeans

sns.set_style('darkgrid')
# palette
plt.rcParams['font.family'] = 'sans-serif'  # Win
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False

# obtain the sheets

SHEET = 0
# load Excel and read sheet
FileName = '../xls/HRB_151318.xlsx'
workbook = load_workbook(FileName)

# Loading the datasets
CheckInData = pd.read_excel(FileName,
                            sheet_name=workbook.sheetnames[SHEET],
                            dtype={'business_id': str})

# Business_Time_Distance:mark "SUB"
subtractions = [0, ]
# second index to last index
for index in range(CheckInData.shape[0]):
    if index < CheckInData.shape[0] - 1:
        subtraction = pd.to_datetime(CheckInData.loc[index + 1, 'business_time']) \
                      - pd.to_datetime(CheckInData.loc[index, 'business_time'])
        subtractions.append(round(subtraction.total_seconds() % 60, 3))
        # print(round(subtraction.total_seconds() % 60, 3))
    else:
        pass


CheckInData['sub'] = subtractions

picture = plt.figure(figsize=(8, 10))
# picture = plt.figure()
ax1 = picture.add_subplot(2, 1, 1)
plt.scatter(y=subtractions,
            x=range(0, CheckInData.shape[0]),
            c='Blue',
            s=15,
            marker='o',
            alpha=1,
            linewidths=0.5,
            edgecolors='white'
            )
plt.title('以时间间隔直角坐标系可视化效果', fontsize=12)
plt.xlabel('按钮顺序编号', fontsize=12)
plt.ylabel('时间间隔', fontsize=12)
# plt.xlim(0, len(subtractions))
# plt.ylim(0, round(max(subtractions), 0) + 1)
# plt.savefig('fig/distance.pdf')
# plt.show()

N_CLUSTER = 5
Cluster_classifier = KMeans(n_clusters=N_CLUSTER,
                            init='k-means++',
                            n_init=10,
                            max_iter=300,
                            tol=0.0001,
                            random_state=0,
                            )


features = np.array([[i, j] for i, j in enumerate(subtractions)])
y_predict = Cluster_classifier.fit_predict(features[:, -1].reshape(-1, 1))


ax2 = picture.add_subplot(2, 1, 2)
plt.scatter(features[y_predict == 0, 0],
            features[y_predict == 0, 1],
            c='Blue',
            s=15,
            marker='o',
            alpha=1,
            label='cluster-1'
            )
plt.scatter(features[y_predict == 1, 0],
            features[y_predict == 1, 1],
            c='Grey',
            s=15,
            marker='h',
            alpha=1,
            label='cluster-2'
            )
plt.scatter(features[y_predict == 2, 0],
            features[y_predict == 2, 1],
            c='Green',
            s=15,
            marker='s',
            alpha=1,
            label='cluster-3'
            )
plt.scatter(features[y_predict == 3, 0],
            features[y_predict == 3, 1],
            c='Orange',
            s=15,
            marker='v',
            alpha=1,
            label='cluster-4'
            )
plt.scatter(features[y_predict == 4, 0],
            features[y_predict == 4, 1],
            c='Purple',
            s=15,
            marker='H',
            alpha=1,
            label='cluster-5'
            )

# plt.scatter(Cluster_classifier.cluster_centers_[:, 0],
#             Cluster_classifier.cluster_centers_[:, 1],
#             s=50,
#             c='red',
#             marker='*',
#             label='centroids'
#             )

plt.title('以时间间隔维度聚类可视化效果', fontsize=12)
plt.xlabel('按钮顺序编号', fontsize=12)
plt.ylabel('时间间隔', fontsize=12)
plt.legend()
# plt.xlim(0, len(subtractions))
# plt.ylim(0, round(max(subtractions), 0) + 1)
plt.savefig('fig/TimeCluster.pdf')
plt.show()
